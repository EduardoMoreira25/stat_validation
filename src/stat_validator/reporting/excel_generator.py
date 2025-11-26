"""
Excel Report Generator Module

Generates Excel reports with summary and daily breakdown sheets.
"""

import logging
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generator for Excel validation reports."""

    def __init__(self):
        """Initialize the Excel generator."""
        pass

    def generate_validation_report(
        self,
        output_path: Path,
        results: List[Any],
        daily_breakdowns: Dict[str, Dict[str, Any]],
        year: int,
        month: int
    ) -> None:
        """
        Generate Excel report with summary and daily breakdown sheets.

        Args:
            output_path: Path to save the Excel file
            results: List of ComparisonResult objects
            daily_breakdowns: Dictionary mapping table_name -> daily breakdown data
            year: Validation year
            month: Validation month
        """
        logger.info(f"Generating Excel report: {output_path}")

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add summary sheet
        self._add_summary_sheet(wb, results, year, month)

        # Add daily breakdown sheets for each table (only if table has data)
        sheets_added = 0
        for result in results:
            table_name = result.table_name
            # Only add sheet if table has non-zero row counts
            if table_name in daily_breakdowns and (result.sap_stats.row_count > 0 or result.dremio_stats.row_count > 0):
                self._add_daily_breakdown_sheet(
                    wb,
                    table_name,
                    daily_breakdowns[table_name]
                )
                sheets_added += 1

        logger.info(f"Added {sheets_added} daily breakdown sheets (tables with data)")

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved: {output_path}")

    def _add_summary_sheet(
        self,
        wb: Workbook,
        results: List[Any],
        year: int,
        month: int
    ) -> None:
        """Add summary sheet with validation results."""
        ws = wb.create_sheet("Summary", 0)

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title
        ws['A1'] = "DBT MONTHLY VALIDATION SUMMARY"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Add metadata
        ws['A2'] = f"Period: {year}-{month:02d}"
        ws['A3'] = f"Total tables: {len(results)}"
        ws['A4'] = f"Passed: {sum(1 for r in results if r.overall_status == 'PASS')}"
        ws['A5'] = f"Failed: {sum(1 for r in results if r.overall_status == 'FAIL')}"

        # Add table headers (starting at row 7)
        headers = ['Table', 'Status', 'SAP Rows', 'Dremio Rows', 'Row Diff %']
        ws.append([])  # Empty row
        header_row = 7
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Sort results by absolute row difference percentage (largest first)
        sorted_results = sorted(results, key=lambda r: abs(r.row_count_diff_pct), reverse=True)

        # Add data rows
        for result in sorted_results:
            row = [
                result.table_name,
                result.overall_status,
                result.sap_stats.row_count,
                result.dremio_stats.row_count,
                round(result.row_count_diff_pct, 2)
            ]
            ws.append(row)

            # Apply styling to the last row
            current_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border

                # Apply status-based fill
                if col_idx == 2:  # Status column
                    if result.overall_status == 'PASS':
                        cell.fill = pass_fill
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = fail_fill
                        cell.font = Font(color="9C0006")

                # Right-align numeric columns
                if col_idx >= 3:
                    cell.alignment = Alignment(horizontal='right')

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15

        # Freeze header row
        ws.freeze_panes = ws['A8']

    def _add_daily_breakdown_sheet(
        self,
        wb: Workbook,
        table_name: str,
        breakdown_data: Dict[str, Any]
    ) -> None:
        """Add daily breakdown sheet for a table."""
        # Sanitize sheet name (Excel has 31 char limit and doesn't allow certain chars)
        sheet_name = self._sanitize_sheet_name(table_name)

        try:
            ws = wb.create_sheet(sheet_name)
        except Exception as e:
            logger.warning(f"Could not create sheet for {table_name}: {e}")
            return

        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add title
        ws['A1'] = f"Daily Breakdown: {table_name}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.merge_cells('A1:C1')

        ws['A2'] = f"Period: {breakdown_data['year']}-{breakdown_data['month']:02d}"

        # Add headers (starting at row 4)
        headers = ['Date', 'SAP Count', 'Dremio Count']
        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add daily data
        daily_data = breakdown_data.get('daily_data', [])
        for data_row in daily_data:
            # Convert date to string if it's a datetime object
            date_val = data_row['date']
            if isinstance(date_val, pd.Timestamp):
                date_str = date_val.strftime('%Y-%m-%d')
            elif isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
            else:
                date_str = str(date_val).split(' ')[0]  # Take just the date part

            row = [
                date_str,
                int(data_row.get('sap_count', 0)),
                int(data_row.get('dremio_count', 0))
            ]
            ws.append(row)

            # Apply styling to the last row
            current_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border

                # Right-align numeric columns
                if col_idx >= 2:
                    cell.alignment = Alignment(horizontal='right')

        # Add total row
        if daily_data:
            ws.append([])  # Empty row
            total_row = ws.max_row + 1
            ws.cell(row=total_row, column=1, value="TOTAL")
            ws.cell(row=total_row, column=1).font = Font(bold=True)

            # Calculate totals
            sap_total = sum(int(row.get('sap_count', 0)) for row in daily_data)
            dremio_total = sum(int(row.get('dremio_count', 0)) for row in daily_data)

            ws.cell(row=total_row, column=2, value=sap_total)
            ws.cell(row=total_row, column=2).font = Font(bold=True)
            ws.cell(row=total_row, column=2).alignment = Alignment(horizontal='right')

            ws.cell(row=total_row, column=3, value=dremio_total)
            ws.cell(row=total_row, column=3).font = Font(bold=True)
            ws.cell(row=total_row, column=3).alignment = Alignment(horizontal='right')

            # Add borders to total row
            for col_idx in range(1, 4):
                ws.cell(row=total_row, column=col_idx).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        # Freeze header row
        ws.freeze_panes = ws['A5']

    def _sanitize_sheet_name(self, name: str) -> str:
        """
        Sanitize sheet name to comply with Excel rules.

        Excel sheet names:
        - Cannot exceed 31 characters
        - Cannot contain: \ / ? * [ ]
        - Cannot be empty
        """
        # Remove invalid characters
        invalid_chars = ['\\', '/', '?', '*', '[', ']']
        sanitized = name
        for char in invalid_chars:
            sanitized = sanitized.replace(char, '_')

        # Truncate to 31 characters
        if len(sanitized) > 31:
            sanitized = sanitized[:31]

        return sanitized
